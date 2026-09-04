from __future__ import annotations

import re
from collections import Counter
from pathlib import Path

import openpyxl


WORKBOOK = Path(r"C:\Users\JesusLovesMe\Documents\CivPro_tag.xlsx")
OUTLINE = Path(r"C:\FOC\Workspace\OUTLINE_CODES_COMPLETE.md")


def load_valid_civpro_codes() -> dict[str, str]:
    text = OUTLINE.read_text(encoding="utf-8")
    start = text.index("# CIVIL PROCEDURE")
    end = text.index("END OF REFERENCE", start)
    section = text[start:end]
    codes: dict[str, str] = {}
    for line in section.splitlines():
        match = re.match(r"\s*(9\d{7})\s+(.+?)\s*$", line)
        if match:
            codes[match.group(1)] = match.group(2)
    return codes


def norm(*values: object) -> str:
    return " ".join("" if v is None else str(v) for v in values).lower()


# These are rows where the official explanation points cleanly to a doctrine
# that broad keyword rules can misread because several Civil Procedure areas
# appear in the same explanation.
MANUAL_BY_QID = {
    14013: "93110101",
    14017: "93110101",
    14003: "93110300",
    14018: "93110300",
    14012: "93110100",
    14008: "93110300",
    14016: "93110200",
    14007: "93110300",
    14000: "93110300",
    14006: "93110300",
    14010: "93110101",
    14014: "93110101",
    20445: "91090300",
    19760: "93110101",
    18537: "91090200",
    18847: "94010100",
    18831: "94010200",
    21297: "94010000",
    20060: "95070303",
    17123: "94010100",
    20718: "94010100",
    17526: "95070103",
    20648: "94020201",
    20281: "91090400",
    14035: "94030100",
    14052: "94050200",
    14032: "94050300",
    14027: "94040300",
    14132: "92080104",
    14123: "92080201",
    14138: "92080105",
    14141: "92080105",
    14125: "92080105",
    14137: "92080202",
    14122: "92080101",
    14109: "92080202",
    14129: "92080105",
    14120: "92080105",
    14121: "92080104",
    14117: "92080202",
    14139: "92080204",
    14116: "92080105",
    14124: "92080104",
    14107: "92080204",
    14108: "92080201",
    14104: "92080104",
    14111: "92080204",
    14131: "92080202",
    14142: "92080201",
    14113: "92080104",
    14099: "92080104",
    14136: "92080202",
    14145: "92080104",
    14112: "92080203",
    14118: "92080105",
    14135: "92080204",
    14133: "92080201",
    14102: "92080204",
    14106: "92080201",
    14096: "92080203",
    14098: "92080104",
    14103: "92080204",
    14126: "92080204",
    14130: "92080202",
    14127: "92080204",
    14114: "92080105",
    14144: "92080103",
    14097: "92080203",
    14115: "92080203",
    14173: "95070402",
    14169: "95070302",
    14182: "95070302",
    14165: "95070302",
    14167: "95070200",
    14172: "95070302",
    14190: "95070302",
    14157: "95070302",
    14179: "95070303",
    14193: "95070303",
    14176: "95070102",
    14194: "95070705",
    14201: "95070103",
    14155: "95070407",
    14164: "95070302",
    14158: "95070302",
    14177: "unsure",
    14198: "95070705",
    14184: "95070101",
    14147: "95070600",
    14154: "95070303",
    14150: "95070302",
    14189: "95070302",
    14082: "96060100",
    14083: "unsure",
    14085: "96060100",
    14081: "96060100",
    14088: "96060100",
    14205: "97100200",
    14210: "97100300",
    14211: "97100300",
    14218: "97100300",
    14202: "97100300",
    14203: "97100300",
    14217: "97100300",
    14207: "97100300",
    14204: "97100300",
    14220: "97100300",
    14221: "97100100",
    14213: "97100100",
    14216: "97100300",
    21253: "97100400",
    20755: "97100300",
    20829: "94030100",
    18510: "94040300",
    20086: "94010000",
    21106: "94040100",
    17763: "94020302",
    18704: "94020302",
    19710: "94040100",
    17226: "94020303",
    18860: "95070302",
    19525: "94030100",
    18282: "91090200",
    19296: "94010000",
    18388: "95070101",
    18531: "91090100",
    18976: "94010200",
    20589: "94050100",
    20614: "94050200",
    17908: "94050200",
    20574: "94030100",
    20029: "97100300",
    18979: "96060100",
    19098: "95070500",
    19721: "92080104",
    20621: "92080104",
    19578: "95070600",
    17739: "95070302",
    17429: "95070402",
    17381: "95070402",
    20302: "95070404",
    18648: "94040200",
    17096: "97100100",
    20562: "92080105",
    19813: "95070101",
    17023: "95070403",
    19398: "96060200",
    21353: "92080205",
    18817: "96060100",
    17103: "96060100",
    17174: "96060100",
    21602: "95070406",
    17868: "94010100",
    18294: "95070104",
    17687: "92080201",
    19022: "95070402",
    20505: "95070705",
    20486: "95070200",
    21422: "95070402",
    21479: "94010300",
    18894: "93110300",
    17384: "97100300",
}


def classify(qid: int, question: str, explanation: str, topic: str) -> str:
    if qid in MANUAL_BY_QID:
        return MANUAL_BY_QID[qid]

    t = norm(question, explanation, topic)

    if "jury" in t:
        if "demand" in t or "waiver" in t or "withdraw" in t:
            return "91090200"
        if "voir dire" in t or "peremptory" in t or "challenge" in t or "juror" in t:
            return "91090300"
        if "instruction" in t:
            return "91090400"
        return "91090100"

    if "appeal" in topic.lower() or "standard of review" in t or "de novo" in t or "clearly erroneous" in t or "abuse of discretion" in t:
        if "standard of review" in t or "de novo" in t or "clearly erroneous" in t or "abuse of discretion" in t or "harmless error" in t:
            return "93110300"
        if "interlocutory" in t or "collateral order" in t or "mandamus" in t or "injunction" in t:
            return "93110100"
        if "final judgment rule" in t or "final decision" in t or "final order" in t:
            return "93110101"
        return "93110200"

    if "law applied" in topic.lower() or "erie" in t or "hanna" in t or "klaxon" in t or "choice-of-law" in t or "choice of law" in t:
        if "federal common law" in t or "clearfield" in t:
            return "96060200"
        return "96060100"

    if "verdict" in topic.lower() or "judgment" in topic.lower() or "preclusion" in t or "res judicata" in t or "collateral estoppel" in t:
        if "preclusion" in t or "res judicata" in t or "collateral estoppel" in t or "full faith and credit" in t:
            return "97100300"
        if "default" in t:
            return "97100400"
        if "dismissal" in t or "dismissals" in t:
            return "97100500"
        if "judicial finding" in t or "bench trial" in t:
            return "97100200"
        return "97100100"

    if "jurisdiction" in topic.lower() or "venue" in topic.lower() or "remove" in t or "remand" in t or "personal jurisdiction" in t:
        if "diversity" in t or "citizen" in t or "amount in controversy" in t or "domicile" in t or "1332" in t:
            return "94010100"
        if "federal question" in t or "arising under" in t or "1331" in t:
            return "94010200"
        if "supplemental" in t or "pendent" in t or "same case or controversy" in t or "common nucleus" in t:
            return "94010300"
        if "remove" in t or "removal" in t or "1441" in t or "1446" in t:
            return "94030100"
        if "remand" in t:
            return "94030200"
        if "forum non conveniens" in t:
            return "94050300"
        if "transfer" in t or "1404" in t or "1406" in t:
            return "94050200"
        if "venue" in t or "1391" in t:
            return "94050100"
        if "waived formal service" in t or "waiver of service" in t:
            return "94040300"
        if "service of process" in t or "summons" in t or "rule 4" in t:
            return "94040200"
        if "notice" in t or "due process notice" in t or "mullane" in t or "jones v. flowers" in t:
            return "94040100"
        if "consent" in t:
            return "94020201"
        if "domicile" in t:
            return "94020203"
        if "long arm" in t or "long-arm" in t:
            return "94020301"
        if "minimum contacts" in t or "purposeful avail" in t:
            return "94020302"
        if "substantial business" in t or "market" in t:
            return "94020303"
        if "personal jurisdiction" in t:
            return "94020100"
        return "94010000"

    if "motion" in topic.lower() or "rule 12" in t or "summary judgment" in t or "jmol" in t or "judgment as a matter of law" in t:
        if "judgment on the pleadings" in t:
            return "92080101"
        if "more definite statement" in t:
            return "92080102"
        if "motion to strike" in t:
            return "92080103"
        if "dismiss" in t or "12(b)" in t or "failure to state" in t:
            return "92080104"
        if "summary judgment" in t or "rule 56" in t:
            return "92080105"
        if "renewed" in t or "jnov" in t:
            return "92080202"
        if "judgment as a matter of law" in t or "rule 50" in t or "jmol" in t:
            return "92080201"
        if "relief from judgment" in t or "rule 60" in t:
            return "92080203"
        if "new trial" in t or "rule 59" in t:
            return "92080204"
        if "remittitur" in t or "additur" in t:
            return "92080205"
        return "92080000"

    if "pretrial" in topic.lower() or "pleading" in t or "discovery" in t or "class action" in t or "rule 11" in t:
        if "amend" in t or "relation back" in t or "rule 15" in t:
            return "95070200"
        if "class action" in t or "rule 23" in t:
            return "95070303"
        if "interpleader" in t or "intervention" in t or "joinder" in t or "indispensable" in t or "necessary party" in t:
            return "95070302"
        if "counterclaim" in t:
            return "95070103"
        if "crossclaim" in t:
            return "95070104"
        if "answer" in t and "complaint" in t:
            return "95070102"
        if "complaint" in t or "plausible" in t or "rule 8" in t or "rule 9" in t:
            return "95070101"
        if "inadvertent disclosure" in t:
            return "95070406"
        if "work product" in t or "privilege" in t:
            return "95070405"
        if "e-discovery" in t or "electronically stored" in t or "spoliation" in t:
            return "95070404"
        if "sanction" in t or "rule 37" in t:
            return "95070407"
        if "interrogator" in t or "deposition" in t or "subpoena" in t or "request for production" in t:
            return "95070402"
        if "scope of discovery" in t or "relevant to any party" in t or "proportional" in t:
            return "95070403"
        if "initial disclosure" in t:
            return "95070401"
        if "rule 11" in t:
            return "95070500"
        if "pretrial conference" in t or "rule 16" in t:
            return "95070600"
        if "voluntary dismissal" in t:
            return "95070701"
        if "involuntary dismissal" in t:
            return "95070702"
        if "summary judgment" in t:
            return "95070703"
        if "default judgment" in t:
            return "95070704"
        if "temporary restraining order" in t or "preliminary injunction" in t:
            return "95070705"
        return "95070000"

    return "unsure"


def iter_rows():
    wb = openpyxl.load_workbook(WORKBOOK)
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    cols = {h: i + 1 for i, h in enumerate(headers)}
    for row in range(2, ws.max_row + 1):
        yield {
            "row": row,
            "qid": int(ws.cell(row, cols["BARMATRIX Q#"]).value),
            "question": ws.cell(row, cols["Question"]).value or "",
            "explanation": ws.cell(row, cols["Answer Explanation"]).value or "",
            "topic": ws.cell(row, cols["topic"]).value or "",
        }


def main(write: bool = False) -> None:
    valid_codes = load_valid_civpro_codes()
    wb = openpyxl.load_workbook(WORKBOOK)
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    cols = {h: i + 1 for i, h in enumerate(headers)}

    assigned = []
    bad = []
    for item in iter_rows():
        code = classify(item["qid"], item["question"], item["explanation"], item["topic"])
        if code != "unsure" and code not in valid_codes:
            bad.append((item["row"], item["qid"], code))
        assigned.append((item, code))
        if write:
            ws.cell(item["row"], cols["Outline_code"]).value = code

    counts = Counter(code for _, code in assigned)
    print(f"rows={len(assigned)}")
    print(f"unique_codes={len(counts)}")
    print(f"unsure={counts.get('unsure', 0)}")
    print(f"invalid_codes={len(bad)}")
    for code, count in counts.most_common():
        label = valid_codes.get(code, "")
        print(f"{count:3d} {code:8s} {label}")
    if bad:
        print("BAD", bad)
    if counts.get("unsure", 0):
        print("UNSURE_ROWS")
        for item, code in assigned:
            if code == "unsure":
                print(item["row"], item["qid"], item["topic"], item["question"][:120])

    if write:
        wb.save(WORKBOOK)

    return 1 if bad else 0


if __name__ == "__main__":
    import sys

    sys.exit(main(write="--write" in sys.argv))
