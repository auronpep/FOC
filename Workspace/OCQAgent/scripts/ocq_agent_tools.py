#!/usr/bin/env python3
"""Deterministic tooling for the OCQ isolated-agent workflow."""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import os
import re
import sys
import time
from collections import Counter, defaultdict
from contextlib import contextmanager
from datetime import datetime, timezone
from pathlib import Path
from typing import Iterable

import openpyxl

QUESTION_HEADERS = [
    "BID",
    "Question",
    "A",
    "B",
    "C",
    "D",
]

ANSWER_COLUMNS = [
    "agent_id",
    "bid",
    "answer_choice",
    "confidence_label",
    "timestamp",
]

ROSTER_COLUMNS = [
    "agent_id",
    "display_name",
    "primary_model",
    "soul_text",
    "enabled",
    "notes",
]

VALID_CHOICES = {"A", "B", "C", "D"}
VALID_CONFIDENCE_LABELS = {"easy", "medium", "hard"}


def utc_now() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def normalize_bid(value: object) -> str:
    if value is None:
        raise ValueError("blank BID")
    text = str(value).strip()
    if text.endswith(".0") and text[:-2].isdigit():
        text = text[:-2]
    if not re.fullmatch(r"\d+", text):
        raise ValueError(f"invalid BID: {value!r}")
    return text


def normalize_agent_id(value: str) -> str:
    agent_id = value.strip().lower()
    if not re.fullmatch(r"[a-z][a-z0-9-]{1,40}", agent_id):
        raise ValueError(f"invalid agent_id {value!r}; use lowercase letters, digits, and hyphens")
    return agent_id


def ensure_parent(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)


def sha256_text(text: str) -> str:
    return hashlib.sha256(text.encode("utf-8")).hexdigest()


def clean_cell(value: object) -> str:
    if value is None:
        return ""
    return str(value).replace("\r\n", "\n").replace("\r", "\n").strip()


def single_line_cell(value: object) -> str:
    return re.sub(r"\s+", " ", clean_cell(value)).strip()


def workbook_rows(workbook_path: Path) -> list[dict[str, str]]:
    if not workbook_path.exists():
        raise FileNotFoundError(f"workbook not found: {workbook_path}")

    wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        ws = wb["QuestionBank"] if "QuestionBank" in wb.sheetnames else wb.active
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        if header is None:
            raise ValueError("QuestionBank sheet is empty")
        headers = [clean_cell(value) for value in header]
        if headers != QUESTION_HEADERS:
            raise ValueError(f"unexpected headers: {headers}")

        records: list[dict[str, str]] = []
        seen: set[str] = set()
        for row_number, row in enumerate(rows, start=2):
            if row is None or all(value is None or clean_cell(value) == "" for value in row):
                continue
            record = {name: single_line_cell(row[index] if index < len(row) else None) for index, name in enumerate(headers)}
            bid = normalize_bid(record["BID"])
            if bid in seen:
                raise ValueError(f"duplicate BID {bid} at row {row_number}")
            seen.add(bid)
            record["BID"] = bid
            blank = [name for name in QUESTION_HEADERS if record[name] == ""]
            if blank:
                raise ValueError(f"row {row_number} has blank required fields: {', '.join(blank)}")
            records.append(record)
    finally:
        wb.close()

    if not records:
        raise ValueError("QuestionBank has no data rows")
    return records


def render_question_markdown(record: dict[str, str]) -> str:
    bid = record["BID"]
    question = single_line_cell(record["Question"])
    choice_a = single_line_cell(strip_choice_prefix(record["A"], "A"))
    choice_b = single_line_cell(strip_choice_prefix(record["B"], "B"))
    choice_c = single_line_cell(strip_choice_prefix(record["C"], "C"))
    choice_d = single_line_cell(strip_choice_prefix(record["D"], "D"))
    return (
        "Carefully review and answer this tricky California Bar Exam Multiple Choice Question. "
        "In your response, give me the question number, your answer selection, your confidence level "
        "(easy, medium, hard). Pick exactly one answer choice; do not give a second guess. "
        f"Here is Question Number {bid}. {question} "
        f"Answer Choices: A. {choice_a} B. {choice_b} C. {choice_c} D. {choice_d}"
    )


def strip_choice_prefix(text: str, letter: str) -> str:
    return re.sub(rf"^\s*{letter}\s*[:.)-]\s*", "", text.strip(), flags=re.IGNORECASE)


def export_questions(workbook_path: Path, output_dir: Path) -> dict[str, object]:
    rows = workbook_rows(workbook_path)
    output_dir.mkdir(parents=True, exist_ok=True)
    written = 0
    for record in rows:
        text = render_question_markdown(record)
        (output_dir / f"{record['BID']}.md").write_text(text, encoding="utf-8", newline="\n")
        written += 1
    return {"written": written, "output_dir": str(output_dir)}


def read_ids(path: Path) -> list[str]:
    if not path.exists():
        return []
    ids: list[str] = []
    for line in path.read_text(encoding="utf-8").splitlines():
        text = line.strip()
        if not text or text.startswith("#"):
            continue
        ids.append(normalize_bid(text))
    return ids


def ensure_answer_csv(path: Path) -> None:
    if path.exists():
        return
    ensure_parent(path)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=ANSWER_COLUMNS)
        writer.writeheader()


def answered_ids(answers_path: Path) -> set[str]:
    if not answers_path.exists():
        return set()
    with answers_path.open(newline="", encoding="utf-8") as handle:
        reader = csv.DictReader(handle)
        if reader.fieldnames != ANSWER_COLUMNS:
            raise ValueError(f"unexpected answer CSV header in {answers_path}: {reader.fieldnames}")
        return {normalize_bid(row["bid"]) for row in reader if row.get("bid")}


def prepare_batch(workspace: Path, count: int) -> list[str]:
    if count < 1:
        raise ValueError("count must be at least 1")
    queue_path = workspace / "queue.txt"
    answers_path = workspace / "answers.csv"
    batch_path = workspace / "current_batch.txt"
    ensure_answer_csv(answers_path)
    queue = read_ids(queue_path)
    done = answered_ids(answers_path)
    selected = [bid for bid in queue if bid not in done][:count]
    batch_path.write_text("".join(f"{bid}\n" for bid in selected), encoding="utf-8", newline="\n")
    return selected


def pid_exists(pid: int) -> bool:
    if pid <= 0:
        return False
    try:
        os.kill(pid, 0)
    except OSError:
        return False
    return True


def clear_stale_lock(lock: Path, stale_seconds: int) -> bool:
    try:
        raw_pid = lock.read_text(encoding="ascii", errors="ignore").strip()
        pid = int(raw_pid) if raw_pid else 0
        age = time.time() - lock.stat().st_mtime
    except (FileNotFoundError, OSError, ValueError):
        return False
    if (pid and not pid_exists(pid)) or age > stale_seconds:
        try:
            lock.unlink()
            return True
        except FileNotFoundError:
            return True
        except OSError:
            return False
    return False


@contextmanager
def lock_path(path: Path, timeout_seconds: int = 120, stale_seconds: int = 300):
    lock = path.with_suffix(path.suffix + ".lock")
    deadline = time.time() + timeout_seconds
    fd = None
    while True:
        try:
            fd = os.open(str(lock), os.O_CREAT | os.O_EXCL | os.O_WRONLY)
            os.write(fd, str(os.getpid()).encode("ascii"))
            break
        except FileExistsError:
            if clear_stale_lock(lock, stale_seconds):
                continue
            if time.time() > deadline:
                raise TimeoutError(f"timed out waiting for lock {lock}")
            time.sleep(0.1)
    try:
        yield
    finally:
        if fd is not None:
            os.close(fd)
        try:
            lock.unlink()
        except FileNotFoundError:
            pass


def validate_choice(value: str, field: str) -> str:
    choice = value.strip().upper()
    if choice not in VALID_CHOICES:
        raise ValueError(f"{field} must be A, B, C, or D")
    return choice


def validate_confidence_label(value: str) -> str:
    label = value.strip().lower()
    if label not in VALID_CONFIDENCE_LABELS:
        raise ValueError("confidence_label must be easy, medium, or hard")
    return label


def append_answer(
    workspace: Path,
    agent_id: str,
    bid: str,
    answer_choice: str,
    confidence_label: str,
    timestamp: str | None = None,
) -> dict[str, str]:
    agent_id = normalize_agent_id(agent_id)
    bid = normalize_bid(bid)
    answer_choice = validate_choice(answer_choice, "answer_choice")
    confidence_label = validate_confidence_label(confidence_label)
    timestamp = timestamp or utc_now()
    answers_path = workspace / "answers.csv"
    ensure_answer_csv(answers_path)
    with lock_path(answers_path):
        done = answered_ids(answers_path)
        if bid in done:
            raise ValueError(f"BID {bid} is already present in {answers_path}")
        row = {
            "agent_id": agent_id,
            "bid": bid,
            "answer_choice": answer_choice,
            "confidence_label": confidence_label,
            "timestamp": timestamp,
        }
        with answers_path.open("a", newline="", encoding="utf-8") as handle:
            writer = csv.DictWriter(handle, fieldnames=ANSWER_COLUMNS)
            writer.writerow(row)
    return row


def answer_csv_header() -> str:
    return ",".join(ANSWER_COLUMNS) + "\n"


def render_agent_instructions(
    agent_id: str,
    display_name: str,
    primary_model: str,
    question_bank_root: Path,
) -> str:
    return f"""# {display_name} OCQ Worker

## Role
You are isolated OpenClaw agent `{agent_id}` for blind OCQ answer collection.

## Local Files
- `USER.md`: user preferences. Read it before answering.
- `SOUL.md`: your test-taker personality. Read it before answering.
- `queue.txt`: full assigned question list, one BID per line.
- `current_batch.txt`: the exact BIDs to answer in this run.
- `answers.csv`: your completed answers. Never edit another agent's CSV.

## Shared Question Bank
Read question files from:

`{question_bank_root}`

Each question file is named `<BID>.md` and contains only the blind question fields.

## Workflow
1. Read `USER.md`.
2. Read `SOUL.md`.
3. Read `current_batch.txt`.
4. For each BID in `current_batch.txt`, first confirm it is not already present in `answers.csv`.
5. Read `{question_bank_root}\\<BID>.md`.
6. Answer from the question and answer choices only, while preserving the test-taker personality in `SOUL.md`.
7. Pick exactly one answer choice only. Do not rank choices and do not record a second choice.
8. Assign one confidence label: `easy`, `medium`, or `hard`.
9. Append exactly one row to `answers.csv`, preferably by running `C:\\FOC\\Workspace\\OCQAgent\\scripts\\Add-OCQAnswer.ps1`.
10. Stop after the BIDs in `current_batch.txt`. Do not continue into `queue.txt`.

## Output Rule
Append rows with exactly these CSV columns:

`{",".join(ANSWER_COLUMNS)}`

Do not include a correct answer, answer key, official explanation, second choice, ranked list, or legal research. This is a blind test-taker run.

## Append Helper Example
Run this from PowerShell after deciding:

```powershell
pwsh -NoProfile -File C:\\FOC\\Workspace\\OCQAgent\\scripts\\Add-OCQAnswer.ps1 -AgentId {agent_id} -Bid <BID> -AnswerChoice <A|B|C|D> -ConfidenceLabel <easy|medium|hard>
```

Model selection is controlled by OpenClaw config for this agent. This file does not override it.
"""


def render_soul(display_name: str, primary_model: str, soul_text: str) -> str:
    return f"""# {display_name} Soul

Model selection is controlled by OpenClaw config.

## Test-Taker Profile
{soul_text.strip()}

## Answering Style
- Answer as this test taker, not as a perfect bar tutor.
- Preserve uncertainty.
- If two choices both feel plausible, make the best pick only.
- Use `easy` when the answer feels clear, `medium` when there is some doubt, and `hard` when it is a close call or partial guess.
- Do not rank choices, record a runner-up, or add an explanation to the answer row.
"""


def render_user() -> str:
    return """# USER.md - About Boss

- **Preferred address:** Boss.
- **Project context:** Boss is building FOC/BarMatrix as a practical MBE diagnostic and wrong-answer repair system.
- **Working style:** Boss wants concrete progress, exact files, verified counts, and short status reports.
- **Agent expectation:** Follow the local queue and batch rules exactly. Do not drift into unrequested analysis or open-ended work.
- **OCQ preference:** Do only the counted batch requested. Do not continue beyond `current_batch.txt`.
- **Direct replies:** Address the user as Boss when speaking to them.
"""


def render_identity(display_name: str, primary_model: str) -> str:
    return f"""# IDENTITY.md - Who Am I?

- **Name:** {display_name}
- **Nature:** Isolated OCQ blind-answer worker
- **Model:** Controlled by OpenClaw config.
- **Role:** Answer assigned MBE-style questions as the test-taker described in `SOUL.md`.
- **Human:** Address the user as Boss.
"""


def initialize_agent_workspace(
    agents_root: Path,
    agent_id: str,
    display_name: str,
    primary_model: str,
    soul_text: str,
    question_ids: Iterable[str],
    question_bank_root: Path,
    overwrite_instructions: bool = False,
) -> Path:
    agent_id = normalize_agent_id(agent_id)
    workspace = agents_root / agent_id
    workspace.mkdir(parents=True, exist_ok=True)

    agents_path = workspace / "AGENTS.md"
    user_path = workspace / "USER.md"
    soul_path = workspace / "SOUL.md"
    identity_path = workspace / "IDENTITY.md"
    queue_path = workspace / "queue.txt"
    batch_path = workspace / "current_batch.txt"
    answers_path = workspace / "answers.csv"

    if overwrite_instructions or not agents_path.exists():
        agents_path.write_text(
            render_agent_instructions(agent_id, display_name, primary_model, question_bank_root),
            encoding="utf-8",
            newline="\n",
        )
    if overwrite_instructions or not user_path.exists():
        user_path.write_text(render_user(), encoding="utf-8", newline="\n")
    if overwrite_instructions or not soul_path.exists():
        soul_path.write_text(render_soul(display_name, primary_model, soul_text), encoding="utf-8", newline="\n")
    if overwrite_instructions or not identity_path.exists():
        identity_path.write_text(render_identity(display_name, primary_model), encoding="utf-8", newline="\n")
    ids = [normalize_bid(value) for value in question_ids]
    queue_path.write_text("".join(f"{bid}\n" for bid in ids), encoding="utf-8", newline="\n")
    if not batch_path.exists():
        batch_path.write_text("", encoding="utf-8")
    ensure_answer_csv(answers_path)
    return workspace


def load_roster(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        raise FileNotFoundError(f"roster not found: {path}")
    if path.suffix.lower() == ".csv":
        with path.open(newline="", encoding="utf-8-sig") as handle:
            reader = csv.DictReader(handle)
            if reader.fieldnames != ROSTER_COLUMNS:
                raise ValueError(f"unexpected roster CSV header: {reader.fieldnames}")
            return [{key: clean_cell(row.get(key, "")) for key in ROSTER_COLUMNS} for row in reader]
    if path.suffix.lower() == ".xlsx":
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        try:
            ws = wb.active
            rows = ws.iter_rows(values_only=True)
            header = next(rows, None)
            headers = [clean_cell(value) for value in header or []]
            if headers != ROSTER_COLUMNS:
                raise ValueError(f"unexpected roster workbook header: {headers}")
            records = []
            for row in rows:
                if row is None or all(clean_cell(value) == "" for value in row):
                    continue
                records.append({key: clean_cell(row[index] if index < len(row) else "") for index, key in enumerate(headers)})
            return records
        finally:
            wb.close()
    raise ValueError("roster must be .csv or .xlsx")


def create_roster_template(path: Path) -> dict[str, object]:
    ensure_parent(path)
    examples = [
        ["adam", "OCQ Adam", "openai/gpt-5.3-codex-spark", "", "true", "Bible-name agent. SOUL.md pending from Boss."],
    ]
    if path.suffix.lower() == ".csv":
        with path.open("w", newline="", encoding="utf-8") as handle:
            writer = csv.writer(handle)
            writer.writerow(ROSTER_COLUMNS)
            writer.writerows(examples)
    elif path.suffix.lower() == ".xlsx":
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Agents"
        ws.append(ROSTER_COLUMNS)
        for row in examples:
            ws.append(row)
        for column in ws.columns:
            width = min(max(len(str(cell.value or "")) for cell in column) + 2, 60)
            ws.column_dimensions[column[0].column_letter].width = width
        wb.save(path)
    else:
        raise ValueError("roster template path must end in .csv or .xlsx")
    return {"path": str(path), "rows": len(examples)}


def init_agents_from_roster(
    roster_path: Path,
    agents_root: Path,
    question_ids: Iterable[str],
    question_bank_root: Path,
    overwrite_instructions: bool = False,
) -> dict[str, object]:
    roster = load_roster(roster_path)
    enabled = [row for row in roster if row.get("enabled", "").strip().lower() in {"1", "true", "yes", "y"}]
    created = []
    for row in enabled:
        workspace = initialize_agent_workspace(
            agents_root=agents_root,
            agent_id=row["agent_id"],
            display_name=row["display_name"] or row["agent_id"],
            primary_model=row["primary_model"],
            soul_text=row["soul_text"] or "Default blind MBE test taker.",
            question_ids=question_ids,
            question_bank_root=question_bank_root,
            overwrite_instructions=overwrite_instructions,
        )
        created.append(str(workspace))
    return {"enabled": len(enabled), "workspaces": created}


def summarize_answers(agents_root: Path, output_dir: Path) -> dict[str, object]:
    output_dir.mkdir(parents=True, exist_ok=True)
    all_rows = []
    for answers_path in sorted(agents_root.glob("*/answers.csv")):
        agent_id = answers_path.parent.name
        with answers_path.open(newline="", encoding="utf-8") as handle:
            reader = csv.DictReader(handle)
            if reader.fieldnames != ANSWER_COLUMNS:
                raise ValueError(f"unexpected answer CSV header in {answers_path}: {reader.fieldnames}")
            for row in reader:
                row = dict(row)
                if row.get("agent_id") != agent_id:
                    raise ValueError(f"agent_id mismatch in {answers_path}: {row.get('agent_id')!r} != {agent_id!r}")
                all_rows.append(row)

    all_path = output_dir / "all_answers.csv"
    all_fields = ANSWER_COLUMNS
    with all_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=all_fields)
        writer.writeheader()
        for row in all_rows:
            writer.writerow({field: row.get(field, "") for field in all_fields})

    by_question_path = output_dir / "by_question.csv"
    grouped: dict[str, list[dict[str, str]]] = defaultdict(list)
    for row in all_rows:
        grouped[row["bid"]].append(row)
    with by_question_path.open("w", newline="", encoding="utf-8") as handle:
        fields = ["bid", "response_count", "A", "B", "C", "D", "easy", "medium", "hard", "top_answer"]
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        for bid in sorted(grouped, key=lambda item: int(item)):
            counts = Counter(row["answer_choice"] for row in grouped[bid])
            labels = Counter(row["confidence_label"] for row in grouped[bid])
            top = counts.most_common(1)[0][0] if counts else ""
            writer.writerow(
                {
                    "bid": bid,
                    "response_count": len(grouped[bid]),
                    "A": counts.get("A", 0),
                    "B": counts.get("B", 0),
                    "C": counts.get("C", 0),
                    "D": counts.get("D", 0),
                    "easy": labels.get("easy", 0),
                    "medium": labels.get("medium", 0),
                    "hard": labels.get("hard", 0),
                    "top_answer": top,
                }
            )

    by_agent_path = output_dir / "by_agent.csv"
    agent_counts = Counter(row["agent_id"] for row in all_rows)
    with by_agent_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=["agent_id", "response_count"])
        writer.writeheader()
        for agent_id in sorted(agent_counts):
            writer.writerow({"agent_id": agent_id, "response_count": agent_counts[agent_id]})

    return {"answers": len(all_rows), "all_answers": str(all_path), "by_question": str(by_question_path), "by_agent": str(by_agent_path)}


def parse_question_ids(args: argparse.Namespace) -> list[str]:
    if args.question_ids:
        return [normalize_bid(part) for item in args.question_ids for part in re.split(r"[,\s]+", item) if part.strip()]
    if args.question_file:
        return read_ids(Path(args.question_file))
    if args.workbook:
        return [record["BID"] for record in workbook_rows(Path(args.workbook))]
    raise ValueError("provide --question-ids, --question-file, or --workbook")


def build_openclaw_commands(roster_path: Path, agents_root: Path) -> list[str]:
    commands = []
    for row in load_roster(roster_path):
        if row.get("enabled", "").strip().lower() not in {"1", "true", "yes", "y"}:
            continue
        agent_id = normalize_agent_id(row["agent_id"])
        workspace = agents_root / agent_id
        commands.append(
            "& 'C:\\FOC\\bin\\openclaw.ps1' agents add "
            f"{agent_id} --workspace '{workspace}' --non-interactive"
        )
        commands.append(
            "& 'C:\\FOC\\bin\\openclaw.ps1' agents set-identity "
            f"--agent {agent_id} --name '{row['display_name'] or agent_id}' --workspace '{workspace}' --from-identity"
        )
    return commands


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="OCQ isolated-agent workflow tools")
    sub = parser.add_subparsers(dest="command", required=True)

    p = sub.add_parser("export-questions")
    p.add_argument("--workbook", required=True)
    p.add_argument("--output-dir", required=True)

    p = sub.add_parser("create-roster-template")
    p.add_argument("--path", required=True)

    p = sub.add_parser("init-agents")
    p.add_argument("--roster", required=True)
    p.add_argument("--agents-root", required=True)
    p.add_argument("--question-bank-root", required=True)
    p.add_argument("--workbook")
    p.add_argument("--question-file")
    p.add_argument("--question-ids", nargs="*")
    p.add_argument("--overwrite-instructions", action="store_true")

    p = sub.add_parser("prepare-batch")
    p.add_argument("--workspace", required=True)
    p.add_argument("--count", type=int, required=True)

    p = sub.add_parser("append-answer")
    p.add_argument("--workspace", required=True)
    p.add_argument("--agent-id", required=True)
    p.add_argument("--bid", required=True)
    p.add_argument("--answer-choice", required=True)
    p.add_argument("--confidence-label", required=True)

    p = sub.add_parser("summarize")
    p.add_argument("--agents-root", required=True)
    p.add_argument("--output-dir", required=True)

    p = sub.add_parser("openclaw-commands")
    p.add_argument("--roster", required=True)
    p.add_argument("--agents-root", required=True)

    args = parser.parse_args(argv)
    try:
        if args.command == "export-questions":
            result = export_questions(Path(args.workbook), Path(args.output_dir))
        elif args.command == "create-roster-template":
            result = create_roster_template(Path(args.path))
        elif args.command == "init-agents":
            result = init_agents_from_roster(
                roster_path=Path(args.roster),
                agents_root=Path(args.agents_root),
                question_ids=parse_question_ids(args),
                question_bank_root=Path(args.question_bank_root),
                overwrite_instructions=args.overwrite_instructions,
            )
        elif args.command == "prepare-batch":
            result = {"selected": prepare_batch(Path(args.workspace), args.count)}
        elif args.command == "append-answer":
            result = append_answer(
                workspace=Path(args.workspace),
                agent_id=args.agent_id,
                bid=args.bid,
                answer_choice=args.answer_choice,
                confidence_label=args.confidence_label,
            )
        elif args.command == "summarize":
            result = summarize_answers(Path(args.agents_root), Path(args.output_dir))
        elif args.command == "openclaw-commands":
            result = {"commands": build_openclaw_commands(Path(args.roster), Path(args.agents_root))}
        else:
            raise AssertionError(args.command)
    except Exception as exc:
        print(json.dumps({"ok": False, "error": str(exc)}), file=sys.stderr)
        return 1
    print(json.dumps({"ok": True, "result": result}, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
