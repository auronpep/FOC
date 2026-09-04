from __future__ import annotations

import csv
import os
import re
import shutil
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

TASKS_DIR = Path(__file__).resolve().parent
WORKSPACE_ROOT = TASKS_DIR.parent

WORKBOOK = Path(
    os.environ.get(
        "CONTRACT_WORKBOOK", Path.home() / "Documents" / "contract law tag.xlsx"
    )
)
OUTLINE = WORKSPACE_ROOT / "OUTLINE_CODES_COMPLETE.md"
AUDIT = TASKS_DIR / "contract_law_tag_audit.csv"


def extract_contract_codes() -> dict[str, str]:
    text = OUTLINE.read_text(encoding="utf-8")
    codes: dict[str, str] = {}
    in_valid = False
    for line in text.splitlines():
        if line.strip() == "Valid codes:":
            in_valid = True
            continue
        if not in_valid:
            continue
        if line.startswith("AB=") or line.startswith("#") or line.startswith("="):
            in_valid = False
            continue
        match = re.match(r"\s*(\d{8})\s+(.+?)\s*$", line)
        if match:
            code, path = match.groups()
            if code.startswith("5"):
                codes[code] = path
    return codes


AUDIT_COLUMNS = [
    "excel_row",
    "qid",
    "outline_code",
    "confidence",
    "note",
    "question_excerpt",
    "explanation_excerpt",
]


def norm(value) -> str:
    return "" if value is None else str(value)


RULES: list[tuple[str, str, str, str]] = [
    ("54030400", "Statute of Frauds", "high", r"statute of frauds|writing requirement|in writing|surety|answer for the debt of another|one[- ]year|sale of goods.*\$500|land contract"),
    ("54030700", "Duress", "high", r"duress|improper threat|economic distress|coerc"),
    ("54030800", "Misrepresentation / Fraud", "high", r"misrepresentation|fraud|fraudulent|false statement|conceal|nondisclosure|material fact|scienter|reliance"),
    ("54030600", "Mistake / Lack of Mutual Assent", "high", r"mistake|mutual mistake|unilateral mistake|misunderstanding|lack of mutual assent|no meeting of the minds"),
    ("54030100", "Lack of Capacity", "high", r"minor|infant|incapacity|mental capacity|incompetent|intoxicated"),
    ("54030300", "Unconscionability", "high", r"unconscionab"),
    ("54030200", "Illegality", "high", r"illegal|illegality|licensing requirement|unlicensed|license requirement"),
    ("54030900", "Undue Influence", "high", r"undue influence|dominant party|confidential relationship"),
    ("54031000", "Public Policy", "high", r"public policy|covenant not to compete|noncompete|restraint of trade|exculpatory"),
    ("52040300", "Parol Evidence Rule", "high", r"parol evidence|integrated|integration|prior or contemporaneous|extrinsic evidence"),
    ("52040200", "Interpretation", "high", r"interpretation|ambiguous term|course of performance|course of dealing|usage of trade|construe|meaning of"),
    ("52040100", "Gap-Fillers", "high", r"gap[- ]filler|reasonable price|open price|missing term|supply a term|default term"),
    ("53070100", "Third-Party Beneficiary Contracts", "high", r"third[- ]party beneficiar|intended beneficiar|incidental beneficiar|vested rights"),
    ("53070200", "Assignment of Rights", "high", r"assignment|assignor|assignee|assigned right"),
    ("53070300", "Delegation of Duties", "high", r"delegation|delegatee|delegated duty|personal services"),
    ("51060304", "Specific performance", "high", r"specific performance|unique goods|unique property|inadequate legal remedy"),
    ("51060302", "Rescission", "high", r"rescission|rescind|avoid the contract|cancel the contract"),
    ("51060303", "Reformation", "high", r"reformation|reform the contract"),
    ("51060305", "Injunctions", "high", r"injunction|enjoin"),
    ("51060105", "Liquidated damages", "high", r"liquidated damages|penalty clause"),
    ("51060107", "Duty to mitigate", "high", r"mitigate|avoidable damages|\bcover\b|\bcovered\b|\bcovering\b"),
    ("51060103", "Consequential damages", "high", r"consequential damages|foreseeable damages|hadley|lost profits"),
    ("51060104", "Incidental damages", "high", r"incidental damages"),
    ("51060102", "Reliance damages", "high", r"reliance damages|expenses incurred in reliance"),
    ("51060106", "Nominal damages", "high", r"nominal damages"),
    ("51060101", "Expectation damages", "high", r"expectation damages|benefit of the bargain|lost profits|contract price|market price|cover price|measure of damages"),
    ("51060201", "Legal Restitution", "high", r"restitution|unjust enrichment|reasonable value|quantum meruit"),
    ("55020300", "Promissory Estoppel", "high", r"promissory estoppel|detrimental reliance|reasonably relied|reliance was foreseeable"),
    ("55020500", "Implied-in-law contracts", "high", r"quasi[- ]contract|contract implied[- ]in[- ]law|implied in law"),
    ("55020400", "Implied-in-fact contracts", "high", r"implied[- ]in[- ]fact|implied in fact|conduct.*manifest"),
    ("55020204", "Preexisting duty", "high", r"preexisting duty|already obligated|modification.*consideration|promise to do what"),
    ("55020202", "Illusory promises", "high", r"illusory|unfettered discretion|at will|may cancel at any time"),
    ("55020203", "Requirement and output contracts", "high", r"requirements contract|output contract|exclusive dealing|best efforts"),
    ("55020205", "Sufficiency of consideration", "high", r"adequacy of consideration|sufficiency of consideration|past consideration|moral obligation|peppercorn|grossly inadequate|fair at the time"),
    ("55020201", "Bargain and exchange", "medium", r"consideration|bargain|detriment|benefit|induced|forbearance|abandon.*claim|good faith claim"),
    ("55020101", "Offer", "high", r"offer|revocation|option contract|firm offer|lapse|advertisement|invitation to deal|preliminary negotiation"),
    ("55020102", "Acceptance", "high", r"acceptance|mailbox rule|mirror image|battle of the forms|additional terms|shipment|silence as acceptance"),
    ("55020100", "Mutual Assent", "medium", r"mutual assent|objective manifestation|manifested intent|definiteness|agreement"),
    ("56050200", "Modifications", "high", r"modification|modify the contract"),
    ("56050403", "Anticipatory repudiation", "high", r"anticipatory repudiation|repudiat|adequate assurance|unequivocal"),
    ("56050402", "Perfect Tender Rule", "high", r"perfect tender|nonconforming goods|reject.*goods|cure"),
    ("56050401", "Material v. Minor", "high", r"material breach|minor breach|substantial performance"),
    ("56050102", "Conditions", "high", r"condition precedent|condition subsequent|express condition|constructive condition|satisfaction condition|occurrence of a condition"),
    ("56050101", "Covenants", "medium", r"covenant|promise rather than a condition"),
    ("56050601", "Impossibility", "high", r"impossibility|impossible|death or incapacity|destroyed|destruction of"),
    ("56050602", "Impracticability", "high", r"impracticability|commercially impracticable|extreme and unreasonable difficulty"),
    ("56050603", "Frustration of purpose", "high", r"frustration of purpose|principal purpose.*frustrated"),
    ("56050604", "Waiver", "high", r"waiver|waive"),
    ("56050605", "Estoppel", "high", r"estoppel"),
    ("56050607", "Accord and satisfaction", "high", r"accord and satisfaction|accord|satisfaction"),
    ("56050608", "Novation", "high", r"novation|substituted obligor"),
    ("56050606", "Discharge by subsequent agreement", "medium", r"mutual rescission|subsequent agreement|release"),
    ("56050500", "Discharge of Contracts", "medium", r"discharge|duty discharged"),
    ("56050300", "Performance", "medium", r"performance|tender of performance|installment"),
    ("56050400", "Breach", "medium", r"breach"),
    ("55010101", "Uniform Commercial Code (UCC)", "medium", r"ucc|article 2|sale of goods|merchant"),
    ("55010102", "Common law", "medium", r"common law"),
    ("55010201", "Bilateral contract", "medium", r"bilateral contract|promise for a promise"),
    ("55010202", "Unilateral contract", "medium", r"unilateral contract|promise for performance"),
]

OVERRIDES: dict[str, tuple[str, str]] = {
    "18276": ("55020205", "adequacy/sufficiency of consideration in equitable enforcement"),
    "17481": ("51060101", "buyer market-contract damages"),
    "18078": ("51060101", "UCC breach-of-warranty damages"),
    "19119": ("53070100", "intended donee beneficiary"),
    "20345": ("51060201", "quasi-contract/restitution remedy"),
    "19716": ("55020102", "promise treated as acceptance"),
    "21701": ("53070200", "assignment of right to receive benefit"),
    "21259": ("56050300", "FOB/risk-of-loss performance rule"),
    "21290": ("55020400", "implied-in-fact promise to pay for requested services"),
    "19461": ("56050608", "novation/substitution of party"),
    "20814": ("53070100", "intended beneficiary rights and vesting"),
    "19373": ("55020205", "past consideration/moral-obligation exception"),
    "21437": ("56050102", "personal satisfaction condition"),
    "21032": ("55020102", "accommodation shipment and acceptance"),
    "19300": ("51060201", "return of price paid after nonconforming tender"),
    "19428": ("51060101", "cost-to-complete expectation damages"),
    "19548": ("51060201", "restitution after excused construction performance"),
    "19539": ("51060101", "lost-volume/no-loss seller damages"),
    "21386": ("56050102", "broker commission condition precedent"),
    "21030": ("52040300", "parol evidence complete integration"),
    "20680": ("56050606", "oral mutual rescission/subsequent agreement"),
    "19953": ("56050300", "FOB delivery and risk-of-loss performance"),
    "18849": ("56050300", "UCC inspection/payment performance rule"),
    "19570": ("56050200", "UCC modification without consideration"),
    "18856": ("56050300", "preinspection payment and UCC performance"),
    "21866": ("56050402", "perfect tender acceptance/rejection options"),
    "20688": ("56050300", "acceptance of goods after inspection"),
    "19172": ("55020101", "option/firm offer"),
    "18427": ("56050603", "frustration of purpose"),
    "17521": ("55020102", "mailbox acceptance"),
    "20782": ("55020204", "preexisting duty rule"),
    "20916": ("54030400", "UCC Statute of Frauds memorandum"),
    "19812": ("56050200", "UCC modification not within Statute of Frauds"),
    "18327": ("56050200", "UCC modification no-consideration rule"),
    "21023": ("55010101", "UCC warranty disclaimer/as-is rule"),
    "21028": ("56050300", "UCC cash-payment performance rule"),
    "20370": ("56050402", "acceptance/rejection of nonconforming goods"),
    "17413": ("51060101", "buyer cover/market damages"),
    "19397": ("56050603", "frustration from change in law"),
    "21577": ("53070100", "intended beneficiary enforcement"),
    "18723": ("56050102", "broker commission condition prevention"),
    "17932": ("56050102", "constructive condition of performance"),
    "17069": ("56050102", "satisfaction condition precedent"),
    "19092": ("56050102", "broker commission express conditions"),
    "21182": ("55020100", "offer plus acceptance mutual assent"),
    "19652": ("56050102", "condition precedent to duty"),
    "18090": ("51060303", "reformation for writing mismatch"),
    "17975": ("53070100", "third-party beneficiary modification before vesting"),
    "21275": ("51060105", "liquidated damages clause"),
    "21525": ("54030100", "minor's capacity/disaffirmance"),
    "17448": ("55020300", "promissory estoppel despite SOF/consideration issues"),
    "20556": ("56050403", "prospective inability/anticipatory nonperformance"),
    "20883": ("55010202", "unilateral contract"),
    "19393": ("55020101", "unilateral reward offer and revocation"),
    "17271": ("56050403", "prospective inability to perform"),
    "18636": ("51060101", "no damages where substitute performance same price"),
    "17133": ("56050607", "accord and satisfaction"),
    "20399": ("56050403", "anticipatory breach/remedy before performance date"),
    "20851": ("53070200", "assignment distinguished from third-party beneficiary"),
    "21516": ("53070200", "assignment extinguishes assignor rights"),
    "17931": ("unsure", "insurance-contingency agreement has no close Contract code"),
    "17431": ("55020101", "offer construed from reasonable recipient perspective"),
    "19520": ("55020101", "revocation of offer"),
    "17078": ("55020101", "option/firm offer"),
    "21462": ("55020101", "firm offer/option without consideration"),
    "19561": ("55020101", "option contract keeps offer open"),
    "19123": ("56050606", "mutual rescission"),
    "14375": ("56050200", "no-oral-modification clause and reliance on extras"),
    "14392": ("55020100", "objective manifestations of assent"),
    "14439": ("56050200", "common-law modification and unanticipated circumstances"),
    "14405": ("56050402", "nonconforming accommodation shipment"),
    "14417": ("unsure", "at-will permanent employment has no close Contract code"),
    "14416": ("52040200", "conflicting meanings and contract interpretation"),
    "14432": ("55020300", "detrimental reliance/promissory estoppel"),
    "14454": ("55020102", "manner of acceptance"),
    "14424": ("55020300", "subcontractor bid made irrevocable by reliance"),
    "14404": ("56050200", "accepted contract modification"),
    "14441": ("52040200", "express term controls trade usage"),
    "14445": ("55020102", "battle-of-forms material additional term"),
    "14448": ("55020102", "mailbox acceptance and additional terms"),
    "14422": ("55020102", "battle-of-forms additional arbitration term"),
    "14435": ("56050200", "modification not supported by consideration"),
    "14409": ("56050602", "impracticability and assumed risk"),
    "14431": ("56050200", "modification for unforeseen difficulty"),
    "14446": ("56050200", "UCC good-faith modification"),
    "14423": ("56050102", "promise payable when able"),
    "14420": ("55020102", "shipment as acceptance and breach"),
    "14421": ("55020201", "bargained-for consideration"),
    "19029": ("56050400", "breach excuses further performance and supports damages"),
    "14440": ("55020204", "debtor promise exception to preexisting duty"),
    "21009": ("55020205", "past consideration"),
    "14414": ("55020102", "acceptance by making requested loan"),
    "14410": ("55020204", "preexisting duty owed to third person"),
    "14418": ("56050200", "unilateral modification accepted before lapse"),
    "14443": ("56050200", "oral modification lacks new consideration"),
    "14447": ("55020300", "promissory estoppel reliance not justifying enforcement"),
    "14452": ("55020102", "battle-of-forms material limitation term"),
    "14407": ("55020102", "battle-of-forms material warranty disclaimer"),
    "14470": ("56050601", "rescission/restitution after impossibility"),
    "14484": ("56050102", "constructive condition of exchange"),
    "14482": ("52040100", "UCC gap-filler for place of delivery"),
    "14489": ("56050602", "impracticability of personal-services performance"),
    "14456": ("56050602", "incapacity makes personal performance impracticable"),
    "14478": ("56050402", "installment nonconformity/substantial impairment"),
    "14487": ("56050607", "accord and satisfaction lacks consideration"),
    "14465": ("56050607", "accord and satisfaction by payment-in-full check"),
    "14494": ("56050604", "waiver of delivery-date condition"),
    "14455": ("56050102", "constructive condition requiring complete performance before payment"),
    "14479": ("56050602", "supervening impracticability"),
    "14476": ("56050102", "express condition and implied good faith"),
    "14477": ("56050401", "minor failure not materially affecting performance"),
    "14468": ("56050604", "waiver of written-notice condition"),
    "14463": ("55020203", "requirements contract breach"),
    "14469": ("56050607", "accord and satisfaction supported by disputed claim"),
    "14488": ("56050300", "shipment contract risk of loss"),
    "14475": ("56050102", "express condition and good-faith efforts"),
    "14490": ("56050604", "waiver of constructive condition"),
    "14491": ("56050403", "mere doubt is not anticipatory repudiation"),
    "14471": ("56050604", "waiver of condition"),
    "14510": ("51060101", "seller action for price/damages"),
    "14531": ("51060101", "expectation damages formula"),
    "14516": ("51060101", "expectation damages after cover"),
    "14524": ("51060101", "lost-volume seller damages"),
    "14526": ("51060103", "foreseeability of consequential damages"),
    "14509": ("51060101", "seller resale/market/lost-volume damages"),
    "14525": ("51060101", "lost profits with reasonable certainty"),
    "14508": ("56050403", "repudiation retraction after reliance"),
    "14518": ("51060101", "UCC breach-of-warranty damages"),
    "14545": ("53070100", "incidental beneficiary"),
    "14538": ("53070100", "intended third-party beneficiary"),
    "14539": ("53070100", "third-party beneficiary rights vested by reliance"),
    "14537": ("53070300", "valid delegation of duties"),
}


def primary_explanation(explanation: str) -> str:
    stops = [
        r"\bA\s+A is incorrect\b",
        r"\bB\s+B is incorrect\b",
        r"\bC\s+C is incorrect\b",
        r"\bD\s+D is incorrect\b",
        r"\bA is therefore incorrect\b",
        r"\bB is therefore incorrect\b",
        r"\bC is therefore incorrect\b",
        r"\bD is therefore incorrect\b",
        r"\bA is incorrect\b",
        r"\bB is incorrect\b",
        r"\bC is incorrect\b",
        r"\bD is incorrect\b",
    ]
    first = len(explanation)
    for pattern in stops:
        match = re.search(pattern, explanation, flags=re.I)
        if match:
            first = min(first, match.start())
    return explanation[:first].strip() or explanation


def classify(question: str, explanation: str) -> tuple[str, str, str]:
    primary = primary_explanation(explanation).lower()
    fallback = f"{primary}\n{question}".lower()
    for code, label, confidence, pattern in RULES:
        if re.search(pattern, primary, flags=re.I | re.S):
            return code, confidence, label
    for code, label, confidence, pattern in RULES:
        if re.search(pattern, fallback, flags=re.I | re.S):
            return code, "fallback", label
    return "unsure", "unsure", "no rule trigger"


def build_audit() -> list[dict[str, str]]:
    wb = load_workbook(WORKBOOK, read_only=True, data_only=False)
    ws = wb.active
    headers = [norm(ws.cell(1, c).value).strip() for c in range(1, ws.max_column + 1)]
    col = {h: i + 1 for i, h in enumerate(headers)}
    rows: list[dict[str, str]] = []
    for r in range(2, ws.max_row + 1):
        qid = norm(ws.cell(r, col["BARMATRIX Q#"]).value)
        question = norm(ws.cell(r, col["Question"]).value)
        explanation = norm(ws.cell(r, col["Answer Explanation"]).value)
        if qid in OVERRIDES:
            code, note = OVERRIDES[qid]
            confidence = "reviewed"
        else:
            code, confidence, note = classify(question, explanation)
        rows.append(
            {
                "excel_row": str(r),
                "qid": qid,
                "outline_code": code,
                "confidence": confidence,
                "note": note,
                "question_excerpt": question[:180].replace("\n", " "),
                "explanation_excerpt": explanation[:320].replace("\n", " "),
            }
        )
    return rows


def write_audit(rows: list[dict[str, str]]) -> None:
    with AUDIT.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=AUDIT_COLUMNS)
        writer.writeheader()
        writer.writerows(rows)


def apply_to_workbook(rows: list[dict[str, str]]) -> Path:
    stamp = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H-%M-%S-%fZ")
    backup = WORKBOOK.with_name(f"{WORKBOOK.stem}.backup-{stamp}{WORKBOOK.suffix}")
    shutil.copy2(WORKBOOK, backup)
    wb = load_workbook(WORKBOOK)
    ws = wb.active
    headers = [norm(ws.cell(1, c).value).strip() for c in range(1, ws.max_column + 1)]
    outline_col = headers.index("Outline_code") + 1
    for row in rows:
        ws.cell(int(row["excel_row"]), outline_col).value = row["outline_code"]
    wb.save(WORKBOOK)
    return backup


def verify(rows: list[dict[str, str]], codes: dict[str, str]) -> dict[str, object]:
    wb = load_workbook(WORKBOOK, read_only=True, data_only=False)
    ws = wb.active
    headers = [norm(ws.cell(1, c).value).strip() for c in range(1, ws.max_column + 1)]
    outline_col = headers.index("Outline_code") + 1
    mismatches = []
    invalid = []
    nonblank = 0
    unsure = 0
    for row in rows:
        excel_row = int(row["excel_row"])
        expected = row["outline_code"]
        actual = norm(ws.cell(excel_row, outline_col).value)
        if actual:
            nonblank += 1
        if actual == "unsure":
            unsure += 1
        if actual != expected:
            mismatches.append((excel_row, row["qid"], expected, actual))
        if actual != "unsure" and actual not in codes:
            invalid.append((excel_row, row["qid"], actual))
    return {
        "rows": len(rows),
        "nonblank": nonblank,
        "unsure": unsure,
        "unique_codes": len({r["outline_code"] for r in rows}),
        "mismatches": mismatches,
        "invalid": invalid,
    }


if __name__ == "__main__":
    official = extract_contract_codes()
    rows = build_audit()
    write_audit(rows)
    counts: dict[str, int] = {}
    for row in rows:
        counts[row["outline_code"]] = counts.get(row["outline_code"], 0) + 1
    print(f"audit={AUDIT}")
    print(f"rows={len(rows)} official_contract_codes={len(official)}")
    print("code_counts:")
    for code, count in sorted(counts.items(), key=lambda item: (-item[1], item[0])):
        print(f"  {code}: {count}")
